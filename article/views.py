import mistune
import stripe
import os
import openpyxl
import re
from openpyxl.utils import get_column_letter

from datetime import datetime
from collections import defaultdict
from datetime import timedelta

from article.models import (
    Category,
    ArticleContent,
    Article,
    ArticleStates,
    CategoryType,
    ArticleVote,
    ArticlesToPublish,
    Payment,
    FavoriteCategory,
)
from article.forms import (
    CategoryForm,
    ArticleForm,
    CategorySearchForm,
    ArticleFilterForm,
)
from roles.utils import PermissionEnum
from notification.utils import send_email

from django.http import (
    HttpResponse,
    HttpResponseForbidden,
    JsonResponse,
    HttpResponseBadRequest,
)
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Avg
from django.db.models import Q
from django.utils import timezone
from django.db.models import Count
from django.db.models import Sum
from django.conf import settings
from django.core.exceptions import ObjectDoesNotExist

from user.models import CustomUser
import json
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook


# Configura Stripe con la clave secreta
stripe.api_key = settings.STRIPE_SECRET_KEY


def global_permissions(request):
    """

    Esta vista se encarga de agregar los permisos del usuario al contexto de la plantilla,
    permitiendo que el usuario acceda a funciones y contenido específicos según sus roles y permisos.

    Si el usuario está autenticado, se obtiene la lista de permisos asociados a sus roles.
    En caso contrario, se retorna una lista vacía de permisos para los usuarios no autenticados.

    Args:
        request (HttpRequest): La solicitud HTTP.

    Returns:
        dict: Un diccionario con la lista de permisos bajo la clave 'permisos'.
    """
    if request.user.is_authenticated:
        permissions = [
            permiso.name
            for rol in request.user.roles.all()
            for permiso in rol.permissions.all()
        ]
        return {'permisos': permissions}
    
    # Return an empty list of permissions for anonymous users
    return {'permisos': []}

def home(request):
    """
    Vista que muestra la página de inicio.

    Muestra una lista de artículos, permitiendo filtrar y ordenar los resultados.
    Los usuarios autenticados pueden ver artículos de categorías gratuitas y de
    categorías por suscripción o pago, siempre y cuando hayan realizado el pago
    correspondiente.

    Args:
        request (HttpRequest): La solicitud HTTP.

    Returns:
        HttpResponse: Renderiza la plantilla 'article/home.html'.
    """

    categories = Category.objects.all()

    if not request.user.is_authenticated:
        permissions = []

        permited_categories = [
            category
            for category in categories
            if category.type == CategoryType.FREE.value
        ]

        not_permited_categories = [
            {"name": category.name, "type": category.type}
            for category in categories
            if category.type != CategoryType.FREE.value
        ]
    else:
        user_payments = Payment.objects.filter(user=request.user)

        payment_status_by_category = defaultdict(lambda: None)
        for payment in user_payments:
            payment_status_by_category[payment.category_id] = payment.status

        permissions = [
            permiso.name
            for rol in request.user.roles.all()
            for permiso in rol.permissions.all()
        ]

        permited_categories = [
            category
            for category in categories
            if category.type
            in (CategoryType.FREE.value, CategoryType.SUSCRIPTION.value)
            or Payment.objects.filter(
                category=category, user=request.user, status="completed"
            ).exists()
        ]

        not_permited_categories = [
            {"name": category.name, "type": category.type}
            for category in categories
            if category.type == CategoryType.PAY.value
            and not Payment.objects.filter(
                category=category, user=request.user, status="completed"
            ).exists()
        ]

    permited_categories_ids = [category.id for category in permited_categories]

    if request.user.is_authenticated:
        favorite_categories_ids = FavoriteCategory.objects.filter(
            user=request.user, category_id__in=permited_categories_ids
        ).values_list("category_id", flat=True)
    else:
        favorite_categories_ids = []

    normal_categories_ids = [
        category.id
        for category in permited_categories
        if category.id not in favorite_categories_ids
    ]

    # Filtrar artículos por conjunto
    favorite_articles = Article.objects.filter(
        state=ArticleStates.PUBLISHED.value, category__id__in=favorite_categories_ids
    )

    normal_articles = Article.objects.filter(
        state=ArticleStates.PUBLISHED.value, category__id__in=normal_categories_ids
    )

    featured_articles = Article.objects.filter(
        is_featured=True, state=ArticleStates.PUBLISHED.value, category_id__in=permited_categories_ids
    )

    img_url_reg = r'!\[.*]\((.*\.(?:jpg|jpeg|png|gif|webp|bmp|svg|tiff|ico)).*\)'
    for fa in featured_articles:
        fa_content = ArticleContent.objects.filter(article=fa).last()
        fa.image_url = re.findall(img_url_reg, str(fa_content.body))
        if not fa.image_url:
            fa.image_url = "static/images/noimage.png"
        else:
            fa.image_url = fa.image_url[0]


    # Aplicar los filtros y ordenamiento a ambos conjuntos
    form = ArticleFilterForm(request.GET or None)
    search_query = request.GET.get("search", "")
    order_by = request.GET.get("order_by", "published_at")
    order_direction = request.GET.get("order_direction", "desc")
    time_range = request.GET.get("time_range", "all")
    # Variable para verificar si alguno de los filtros tiene un valor distinto al predeterminado
    filter_active = False   # Falso por defecto
    if form.is_valid():
        selected_tag = form.cleaned_data.get("tags")
        selected_category = form.cleaned_data.get("category")
        selected_category_type = form.cleaned_data.get("category_type")

        if selected_tag:
            filter_active = True
            favorite_articles = favorite_articles.filter(tags__name__in=[selected_tag])
            normal_articles = normal_articles.filter(tags__name__in=[selected_tag])

        if selected_category:
            filter_active = True
            favorite_articles = favorite_articles.filter(category=selected_category)
            normal_articles = normal_articles.filter(category=selected_category)

        if selected_category_type and selected_category_type != "all":
            filter_active = True
            favorite_articles = favorite_articles.filter(
                category__type=selected_category_type
            )
            normal_articles = normal_articles.filter(
                category__type=selected_category_type
            )

    if time_range != "all":
        now = timezone.now()
        time_filters = {
            "1h": now - timedelta(hours=1),
            "24h": now - timedelta(hours=24),
            "7d": now - timedelta(days=7),
            "30d": now - timedelta(days=30),
            "365d": now - timedelta(days=365),
        }
        if time_range in time_filters:
            favorite_articles = favorite_articles.filter(
                published_at__gte=time_filters[time_range]
            )
            normal_articles = normal_articles.filter(
                published_at__gte=time_filters[time_range]
            )

    if search_query:
        filter_active = True
        favorite_articles = favorite_articles.filter(
            Q(title__icontains=search_query)
            | Q(description__icontains=search_query)
            | Q(tags__name__icontains=search_query)
        ).distinct()

        normal_articles = normal_articles.filter(
            Q(title__icontains=search_query)
            | Q(description__icontains=search_query)
            | Q(tags__name__icontains=search_query)
        ).distinct()

    if order_by == "published_at" and order_direction == "desc":
        order_by = "-published_at"
    elif order_direction == "desc":
        order_by = f"-{order_by}"

    favorite_articles = favorite_articles.order_by(order_by)
    normal_articles = normal_articles.order_by(order_by)

    for article in favorite_articles:
        # Calcular la calificación promedio
        ratings = ArticleVote.objects.filter(article=article)
        avg_rating = ratings.aggregate(Avg("rating"))["rating__avg"]

        # Asignar el promedio de calificación como un atributo del artículo
        article.avg_rating = round(avg_rating, 1) if avg_rating is not None else None

    for article in normal_articles:
        # Calcular la calificación promedio
        ratings = ArticleVote.objects.filter(article=article)
        avg_rating = ratings.aggregate(Avg("rating"))["rating__avg"]

        # Asignar el promedio de calificación como un atributo del artículo
        article.avg_rating = round(avg_rating, 1) if avg_rating is not None else None



    
    img_url_reg = r'!\[.*]\((.*\.(?:jpg|jpeg|png|gif|webp|bmp|svg|tiff|ico)).*\)'
    for fa in normal_articles:
        fa_content = ArticleContent.objects.filter(article=fa).last()
        fa.image_url = re.findall(img_url_reg, str(fa_content.body))
        if not fa.image_url:
            fa.image_url = "static/images/noimage.png"
        else:
            fa.image_url = fa.image_url[0]

    for fa in favorite_articles:
        fa_content = ArticleContent.objects.filter(article=fa).last()
        fa.image_url = re.findall(img_url_reg, str(fa_content.body))
        if not fa.image_url:
            fa.image_url = "static/images/noimage.png"
        else:
            fa.image_url = fa.image_url[0]

    for fa in favorite_articles:
        print(fa.image_url)

    # Crear una lista con todos los articulos
    all_articles = favorite_articles.union(normal_articles)
    #all_articles = all_articles.order_by(order_by)

    for fa in all_articles:
        fa_content = ArticleContent.objects.filter(article=fa).last()
        fa.image_url = re.findall(img_url_reg, str(fa_content.body))
        if not fa.image_url:
            fa.image_url = "static/images/noimage.png"
        else:
            fa.image_url = fa.image_url[0]

    for fa in all_articles:
        print(fa.image_url)

    return render(
        request,
        "article/home.html",
        {
            "permisos": permissions,
            "permited_categories": permited_categories,
            "not_permited_categories": not_permited_categories,
            "favorite_articles": favorite_articles,
            "normal_articles": normal_articles,
            "all_articles": all_articles,
            "featured_articles": featured_articles,
            "form": form,
            "search_query": search_query,
            "order_by": order_by,
            "order_direction": order_direction,
            "time_range": time_range,
            "normal_categories": normal_categories_ids,
            "favorite_categories": favorite_categories_ids,
            "filter_active": filter_active,
        },
    )

    


def forbidden(request):
    """
    Vista que muestra la página de acceso prohibido.

    Args:
        request (HttpRequest): La solicitud HTTP.

    Returns:
        HttpResponse: Renderiza la plantilla 'article/forbidden.html'.
    """

    return render(request, "article/forbidden.html")


# =============================================================================
# Article
# =============================================================================


def article_create(request):
    """
    Vista que permite la creación de un artículo.

    Solo los usuarios autenticados y con el permiso `CREAR_ARTICULOS` pueden
    acceder a esta vista. Si no se cumplen las condiciones, se redirige al
    usuario a la página de login o a la página de acceso prohibido.

    Args:
        request (HttpRequest): La solicitud HTTP.

    Returns:
        HttpResponse: Renderiza la plantilla 'article/create_article.html' o redirige.
    """

    if not request.user.is_authenticated:
        return redirect("login")

    if not request.user.tiene_permisos([PermissionEnum.CREAR_ARTICULOS]):
        return redirect("forbidden")

    if request.method == "POST":
        form = ArticleForm(request.POST)
        if form.is_valid():
            article = form.save(commit=False)
            article.autor = request.user
            article.save()
            form.save_m2m()

            ArticleContent.objects.create(
                body=request.POST.get("body"), autor=request.user, article=article
            )

            return redirect("article-detail", pk=article.id)

    return render(request, "article/article_form.html", {"form": ArticleForm})


def article_update(request, pk):
    """
    Vista que permite la actualización de un artículo.

    Solo los usuarios autenticados y con el permiso `EDITAR_ARTICULOS` pueden
    acceder a esta vista. Si no se cumplen las condiciones, se redirige al
    usuario a la página de login o a la página de acceso prohibido.

    Args:
        request (HttpRequest): La solicitud HTTP.
        pk (int): El ID del artículo a actualizar.

    Returns:
        HttpResponse: Renderiza la plantilla 'article/update_article.html' o redirige.
    """

    if not request.user.is_authenticated:
        return redirect("login")

    if not request.user.tiene_permisos(
        [PermissionEnum.EDITAR_ARTICULOS]
    ) and not request.user.tiene_permisos([PermissionEnum.EDITAR_ARTICULOS_BORRADOR]):
        return redirect("forbidden")

    article = get_object_or_404(Article, pk=pk)

    if request.method == "POST":
        form = ArticleForm(request.POST, instance=article)
        if form.is_valid():
            article = form.save(commit=False)
            article.save()
            form.save_m2m()

            ArticleContent.objects.create(
                body=request.POST.get("body"), autor=request.user, article=article
            )

            return redirect("article-detail", pk=article.id)
    else:
        form = ArticleForm(instance=article)

        last_content = ArticleContent.objects.filter(article=article).last()
        if last_content:
            form.initial["body"] = last_content.body

    return render(request, "article/article_form.html", {"form": form})


def article_update_history(request, pk):
    """
    Vista que muestra el historial de versiones de un artículo.

    Solo los usuarios autenticados y con el permiso `VER_HISTORIAL_ARTICULOS`
    pueden acceder a esta vista. Si no se cumplen las condiciones, se redirige
    al usuario a la página de login o a la página de acceso prohibido.

    Args:
        request (HttpRequest): La solicitud HTTP.
        pk (int): El ID del artículo.

    Returns:
        HttpResponse: Renderiza la plantilla 'article/article_update_history.html' o redirige.
    """

    if not request.user.is_authenticated:
        return redirect("login")

    article = get_object_or_404(Article, pk=pk)

    if not (
        request.user.tiene_permisos([PermissionEnum.EDITAR_ARTICULOS])
        or request.user.tiene_permisos([PermissionEnum.EDITAR_ARTICULOS_BORRADOR])
        or request.user == article.autor
    ):
        return redirect("forbidden")

    if request.method == "POST":
        article_id = request.POST.get("article_id")
        article_content_id = request.POST.get("article_content_id")

        if not article_content_id or not article_id:
            return HttpResponse("No se ha encontrado el contenido", status=404)

        article_content = get_object_or_404(ArticleContent, pk=article_content_id)
        article = get_object_or_404(Article, pk=article_id)

        ArticleContent.objects.create(
            body=article_content.body, autor=request.user, article=article
        )

        return redirect("article-detail", pk=article.pk)

    article = get_object_or_404(Article, pk=pk)
    article_contents_ref = ArticleContent.objects.filter(article=article).order_by(
        "-created_at"
    )

    article_contents = [
        {
            "id": article_content.id,
            "autor": article_content.autor,
            "created_at": article_content.created_at,
            "body": mistune.html(article_content.body),
        }
        for article_content in article_contents_ref
    ]

    return render(
        request,
        "article/article_update_history.html",
        {"article": article, "article_contents": article_contents},
    )


def article_list(request):
    """
    Vista que muestra la lista de artículos.

    Solo los usuarios autenticados y con el permiso `VER_ARTICULOS` pueden
    acceder a esta vista. Si no se cumplen las condiciones, se redirige al
    usuario a la página de login o a la página de acceso prohibido.

    Args:
        request (HttpRequest): La solicitud HTTP.

    Returns:
        HttpResponse: Renderiza la plantilla 'article/article_list.html' o redirige.
    """

    if not request.user.is_authenticated:
        return redirect("login")

    if not request.user.tiene_permisos([PermissionEnum.VER_INICIO]):
        return redirect("forbidden")

    can_create = request.user.tiene_permisos([PermissionEnum.CREAR_ARTICULOS])
    can_edit = request.user.tiene_permisos([PermissionEnum.EDITAR_ARTICULOS])
    can_publish = request.user.tiene_permisos([PermissionEnum.MODERAR_ARTICULOS])

    articles = []

    if can_publish or can_edit:
        articles = Article.objects.all()
    elif can_create:
        articles = Article.objects.filter(autor=request.user)

    return render(
        request,
        "article/article_list.html",
        {
            "articles": articles,
            "can_create": can_create,
        },
    )


def article_detail(request, pk):
    """
    Vista que muestra los detalles de un artículo.

    Si el usuario no está autenticado, se le permitirá ver el artículo solo si
    pertenece a una categoría gratuita. Para los usuarios autenticados, también
    se maneja la lógica de 'me gusta', 'no me gusta' y calificación.

    Args:
        request (HttpRequest): La solicitud HTTP.
        pk (int): El ID del artículo a mostrar.

    Returns:
        HttpResponse: Renderiza la plantilla 'article/article_detail.html'.
    """

    # Fetch the article and its content
    article = get_object_or_404(Article, pk=pk)
    article_content = ArticleContent.objects.filter(article=article).last()

    to_publish_date = ArticlesToPublish.objects.filter(article=article).first()

    if not article_content:
        return HttpResponse("No content for this article", status=404)

    # Check if the article's category is "free"
    is_free = article.category.type == CategoryType.FREE.value
    authenticated = request.user.is_authenticated
    shares_number = request.GET.get("shared", "false") == "true"

    # Handle unauthenticated users (unknown users)
    if not authenticated:
        if is_free:
            # Unknown user can only view the article without interactions
            if article.state == ArticleStates.PUBLISHED.value:
                article.views_number += 1
                article.save()

            # Calculate the average rating for the article
            ratings = ArticleVote.objects.filter(article=article)
            avg_rating = ratings.aggregate(Avg("rating"))["rating__avg"]

            # Check if the average rating is None before rounding
            if avg_rating is not None:
                avg_rating = round(avg_rating, 1)

            # Convert article content body using mistune
            article_render_content = mistune.html(article_content.body)

            return render(
                request,
                "article/article_detail.html",
                {
                    "article": article,
                    "article_render_content": article_render_content,
                    "avg_rating": avg_rating,
                    "to_publish_date": None,
                    # Unauthenticated users cannot edit, like, dislike, or publish
                    "can_edit_as_editor": False,
                    "can_edit_as_author": False,
                    "can_edit": False,
                    "can_publish": False,
                    "can_inactivate": False,
                    "is_moderated_category": article.category.is_moderated,
                    "user_vote": None,  # No vote for unauthenticated users
                    "authenticated": authenticated,
                },
            )
        else:
            # If the article is not free, redirect to login
            return redirect("login")

    # Else block for authenticated users
    else:
        # Ensure user has the right permissions
        if not request.user.tiene_permisos([PermissionEnum.VER_INICIO]):
            return redirect("forbidden")

        # Increment view count
        if article.state == ArticleStates.PUBLISHED.value:
            article.views_number += 1
            article.save()

        # Increment shared count
        if shares_number:
            article.shares_number += 1
            article.save()

        # Fetch the user's vote and rating for the article
        user_vote = ArticleVote.objects.filter(
            article=article, user=request.user
        ).first()

        if article.state == ArticleStates.PUBLISHED.value:
            # Handle like/dislike and rating submissions
            if request.method == "POST":
                if "rating" in request.POST:
                    # Rating submission logic
                    rating_value = int(request.POST.get("rating"))
                    if user_vote:
                        user_vote.rating = rating_value
                        user_vote.save()
                    else:
                        ArticleVote.objects.create(
                            user=request.user, article=article, rating=rating_value
                        )

                elif "like" in request.POST or "dislike" in request.POST:
                    # Handle like/dislike submission
                    if "like" in request.POST:
                        if user_vote and user_vote.vote != ArticleVote.LIKE:
                            if user_vote.vote == ArticleVote.DISLIKE:
                                article.dislikes_number -= 1
                            user_vote.vote = ArticleVote.LIKE
                            article.likes_number += 1
                            user_vote.save()
                        elif not user_vote:
                            ArticleVote.objects.create(
                                user=request.user,
                                article=article,
                                vote=ArticleVote.LIKE,
                            )
                            article.likes_number += 1

                    elif "dislike" in request.POST:
                        if user_vote and user_vote.vote != ArticleVote.DISLIKE:
                            if user_vote.vote == ArticleVote.LIKE:
                                article.likes_number -= 1
                            user_vote.vote = ArticleVote.DISLIKE
                            article.dislikes_number += 1
                            user_vote.save()
                        elif not user_vote:
                            ArticleVote.objects.create(
                                user=request.user,
                                article=article,
                                vote=ArticleVote.DISLIKE,
                            )
                            article.dislikes_number += 1

                    article.save()

        # Calculate the average rating for the article
        ratings = ArticleVote.objects.filter(article=article)
        avg_rating = ratings.aggregate(Avg("rating"))["rating__avg"]

        # Check if the average rating is None before rounding
        if avg_rating is not None:
            avg_rating = round(avg_rating, 1)

        # Check if the user is an admin
        is_admin = request.user.roles.filter(name="Administrador").exists()

        # Check if the user is the author of the article
        is_author = request.user.tiene_permisos([PermissionEnum.CREAR_ARTICULOS])

        # Determine if the user can inactivate (either admin or author)
        can_inactivate = is_admin or is_author

        # Determine if the user can edit as an editor or author
        can_edit_as_editor = is_admin or request.user.tiene_permisos(
            [PermissionEnum.EDITAR_ARTICULOS]
        )
        can_edit_as_author = (
            is_author
            or request.user.tiene_permisos([PermissionEnum.EDITAR_ARTICULOS_BORRADOR])
        ) or is_admin

        # General edit permission (either editor or author)
        can_edit = (can_edit_as_editor or can_edit_as_author) and article.state in [
            ArticleStates.DRAFT.value,
            ArticleStates.REVISION.value,
        ]

        # Can publish only if the user has permission to moderate articles
        can_publish = request.user.tiene_permisos([PermissionEnum.MODERAR_ARTICULOS])

        # Check if the category requires moderation
        is_moderated_category = article.category.is_moderated

        # Convert article content body using mistune
        article_render_content = mistune.html(article_content.body)

        favorite_categories = FavoriteCategory.objects.filter(
            user=request.user
        ).values_list("category_id", flat=True)

        return render(
            request,
            "article/article_detail.html",
            {
                "article": article,
                "article_render_content": article_render_content,
                "can_edit_as_editor": can_edit_as_editor,
                "to_publish_date": to_publish_date,
                "can_edit_as_author": can_edit_as_author,
                "can_edit": can_edit,
                "can_publish": can_publish,
                "can_inactivate": can_inactivate,
                "is_moderated_category": is_moderated_category,
                "avg_rating": avg_rating,
                "user_vote": user_vote,  # Pass both vote and rating information to the template
                "shares_number": article.shares_number,
                "authenticated": authenticated,
                "is_author": is_author,
                "is_admin": is_admin,
                "favorite_categories": favorite_categories,
            },
        )


@login_required
def article_to_revision(request, pk):
    """
    Vista que cambia el estado de un artículo a Revisión.


    Args:
        request (HttpRequest): La solicitud HTTP.
        pk (int): El ID del artículo.
    """

    article = get_object_or_404(Article, pk=pk)

    is_admin = request.user.roles.filter(name="Administrador").exists()
    is_autor = request.user.tiene_permisos([PermissionEnum.CREAR_ARTICULOS])
    # is_editor = request.user.tiene_permisos([PermissionEnum.EDITAR_ARTICULOS])
    # is_publisher = request.user.tiene_permisos([PermissionEnum.MODERAR_ARTICULOS])

    # Solo el autor (cuando el estado es DRAFT) o el admin puede cambiar a REVISIÓN
    if (
        is_admin or (is_autor and article.state == ArticleStates.DRAFT.value)
        # or ((is_editor or is_publisher) and article.state == ArticleStates.EDITED.value)
    ):
        article.change_state(ArticleStates.REVISION.value)
        return redirect("article-detail", pk=pk)

    return HttpResponseForbidden("No puedes editar este contenido")


@login_required
def article_to_published(request, pk):
    """
    Vista que cambia el estado de un artículo a Publicado.

    Args:
        request (HttpRequest): La solicitud HTTP.
        pk (int): El ID del artículo.
    """
    article = get_object_or_404(Article, pk=pk)

    is_admin = request.user.roles.filter(name="Administrador").exists()
    can_publish = is_admin or request.user.tiene_permisos(
        [PermissionEnum.MODERAR_ARTICULOS]
    )
    is_moderated = article.category.is_moderated
    is_author = request.user.tiene_permisos([PermissionEnum.CREAR_ARTICULOS])

    # Solo el publicador o el admin puede cambiar a PUBLICADO
    if is_moderated:
        if can_publish and article.state == ArticleStates.EDITED.value:
            article.change_state(ArticleStates.PUBLISHED.value)
            return redirect("article-detail", pk=pk)
    else:
        if is_author or is_admin or can_publish:
            article.change_state(ArticleStates.PUBLISHED.value)
            return redirect("article-detail", pk=pk)

    return HttpResponseForbidden("No puedes editar este contenido")


@login_required
def article_to_publish_schedule(request, pk):
    """
    Vista que programa una publicación de un archivo.

    Args:
        request (HttpRequest): La solicitud HTTP.
        pk (int): El ID del artículo.
    """

    if request.method != "POST":
        return HttpResponseForbidden("No puedes editar este contenido")

    to_publish_date = request.POST.get("to_publish_date")

    article = get_object_or_404(Article, pk=pk)

    ArticlesToPublish.objects.filter(article=article).delete()

    is_admin = request.user.roles.filter(name="Administrador").exists()
    can_publish = is_admin or request.user.tiene_permisos(
        [PermissionEnum.MODERAR_ARTICULOS]
    )
    is_moderated = article.category.is_moderated
    is_author = request.user.tiene_permisos([PermissionEnum.CREAR_ARTICULOS])

    # Solo el publicador o el admin puede cambiar a PUBLICADO
    if is_moderated:
        if can_publish and article.state == ArticleStates.EDITED.value:
            ArticlesToPublish.objects.create(
                article=article, to_publish_at=to_publish_date
            )
            return redirect("article-detail", pk=pk)
    else:
        if is_author or is_admin or can_publish:
            ArticlesToPublish.objects.create(
                article=article, to_publish_at=to_publish_date
            )
            return redirect("article-detail", pk=pk)

    return HttpResponseForbidden("No puedes editar este contenido")


@login_required
def article_to_edited(request, pk):
    """
    Vista que cambia el estado de un artículo a Editado.

    Args:
        request (HttpRequest): La solicitud HTTP.
        pk (int): El ID del artículo.
    """
    article = get_object_or_404(Article, pk=pk)

    is_admin = request.user.roles.filter(name="Administrador").exists()
    is_editor = is_admin or request.user.tiene_permisos(
        [PermissionEnum.EDITAR_ARTICULOS]
    )

    # Solo el editor o el admin pueden cambiar a EDITADO y el estado actual debe ser REVISIÓN
    if is_editor and article.state == ArticleStates.REVISION.value:
        article.change_state(ArticleStates.EDITED.value)
        return redirect("article-detail", pk=pk)

    return HttpResponseForbidden("No puedes editar este contenido")


@login_required
def article_to_draft(request, pk):
    """
    Vista que cambia el estado de un artículo a Borrador.

    Args:
        request (HttpRequest): La solicitud HTTP.
        pk (int): El ID del artículo.
    """
    article = get_object_or_404(Article, pk=pk)

    is_admin = request.user.roles.filter(name="Administrador").exists()
    is_editor = is_admin or request.user.tiene_permisos(
        [PermissionEnum.EDITAR_ARTICULOS]
    )

    # Solo el autor (cuando el estado es REVISION) o el admin puede cambiar a DRAFT
    if is_admin or (is_editor and article.state == ArticleStates.REVISION.value):
        article.change_state(ArticleStates.DRAFT.value)
        return redirect("article-detail", pk=pk)

    return HttpResponseForbidden("No puedes editar este contenido")


@login_required
def article_to_inactive(request, pk):
    """
    Vista que cambia el estado de un artículo a Inactivo.

    Args:
        request (HttpRequest): La solicitud HTTP.
        pk (int): El ID del artículo.
    """
    article = get_object_or_404(Article, pk=pk)

    is_admin = request.user.roles.filter(name="Administrador").exists()
    is_autor = request.user.tiene_permisos([PermissionEnum.CREAR_ARTICULOS])

    # Solo el autor o el admin puede cambiar a INACTIVO
    if is_admin or is_autor:
        article.change_state(ArticleStates.INACTIVE.value)
        return redirect("article-detail", pk=pk)

    return HttpResponseForbidden("No puedes editar este contenido")

def manage_featured_articles(request):
    """
    Vista para que los administradores gestionen los artículos destacados.

    Solo los usuarios autenticados y con el permiso `MODERAR_ARTICULOS` pueden
    acceder a esta vista. Si no se cumplen las condiciones, se redirige al
    usuario a la página de login o a la página de acceso prohibido.

    Args:
        request (HttpRequest): La solicitud HTTP.

    Returns:
        HttpResponse: Renderiza la plantilla 'admin/manage_featured_articles.html' o redirige.
    """

    # Verificar autenticación
    if not request.user.is_authenticated:
        return redirect("login")

    # Verificar permisos
    if not request.user.tiene_permisos([PermissionEnum.MODERAR_ARTICULOS]):
        return redirect("forbidden")

    # Obtener todos los artículos publicados
    articles = Article.objects.filter(state=ArticleStates.PUBLISHED.value)

    # Si el formulario se envía (POST), procesar cambios
    if request.method == "POST":
        article_id = request.POST.get("article_id")
        action = request.POST.get("action")
        article = Article.objects.filter(id=article_id, state=ArticleStates.PUBLISHED.value).first()

        if article:
            if action == "add":
                article.is_featured = True
            elif action == "remove":
                article.is_featured = False
            article.save()

    # Separar los artículos destacados de los no destacados
    featured_articles = articles.filter(is_featured=True)
    non_featured_articles = articles.filter(is_featured=False)

    return render(
        request,
        "article/manage_featured_articles.html",
        {
            "featured_articles": featured_articles,
            "non_featured_articles": non_featured_articles,
        },
    )



# =============================================================================
# Category views
# =============================================================================


def category_list(request):
    """
    Vista que muestra la lista de categorías y permite buscar, filtrar y ordenar resultados.

    Muestra todas las categorías de manera predeterminada y actualiza los resultados según
    la entrada del formulario de búsqueda.

    Args:
        request (HttpRequest): La solicitud HTTP.

    Returns:
        HttpResponse: Renderiza la plantilla 'article/category_list.html' o redirige.
    """

    if not request.user.is_authenticated:
        return redirect("login")

    can_create_categories = request.user.tiene_permisos(
        [PermissionEnum.MANEJAR_CATEGORIAS]
    )

    if not (
        request.user.tiene_permisos([PermissionEnum.MANEJAR_CATEGORIAS])
        or request.user.tiene_permisos([PermissionEnum.VER_CATEGORIAS])
    ):
        return redirect("forbidden")

    form = CategorySearchForm(request.GET or None)
    categories = Category.objects.all()
    favorite_categories = FavoriteCategory.objects.filter(
        user=request.user
    ).values_list("category_id", flat=True)

    user_payments = Payment.objects.filter(user=request.user)

    payment_status_by_category = defaultdict(lambda: None)
    for payment in user_payments:
        payment_status_by_category[payment.category_id] = payment.status


    permited_categories = [
        category
        for category in categories
        if category.type
        in (CategoryType.FREE.value, CategoryType.SUSCRIPTION.value)
        or Payment.objects.filter(
            category=category, user=request.user, status="completed"
        ).exists()
    ]

    not_permited_categories = [
            category
            for category in categories
            if category.type == CategoryType.PAY.value
            and not Payment.objects.filter(
                category=category, user=request.user, status="completed"
            ).exists()
        ]

    if form.is_valid():
        search_term = form.cleaned_data.get("search_term")
        order_by = form.cleaned_data.get("order_by", "name")
        filter_type = form.cleaned_data.get("filter_type", "all")

        # Filtrar por título que contenga el término de búsqueda
        if search_term:
            categories = categories.filter(name__icontains=search_term)

        # Filtrar por tipo de categoría
        if filter_type != "all":
            categories = categories.filter(type=filter_type)

        # Ordenar los resultados
        categories = categories.order_by(order_by)

    return render(
        request,
        "article/category_list.html",
        {
            "form": form,
            "categories": categories,
            "favorite_categories": favorite_categories,
            "can_create_categories": can_create_categories,
            "permited_categories": permited_categories,
            "not_permited_categories": not_permited_categories,
        },
    )


@login_required

def toggle_favorite_category(request, pk):
    """    
    Cambia el estado de favorito de una categoría para el usuario autenticado.
    Esta vista maneja el cambio del estado de favorito de una categoría para el usuario
    que realiza la solicitud. Si la categoría ya está marcada como favorita, se eliminará
    de los favoritos. Si no está marcado como favorito, se agregará a los favoritos.

    Args:
        request (HttpRequest): La solicitud HTTP.
        pk (int): El ID de la categoría.
    Returns:
        JsonResponse: Un objeto JSON con la clave "status" y el valor correspondiente:
            - {"status": "removed"} si la categoría estaba marcada como favorita y se ha eliminado.
            - {"status": "added"} si la categoría no estaba marcada como favorita y se ha agregado.
        HttpResponseBadRequest: Si la solicitud no es válida.
    """    

    if request.method == "POST":
        category = Category.objects.get(id=pk)
        favorite, created = FavoriteCategory.objects.get_or_create(
            user=request.user, category=category
        )

        if not created:
            # Si ya existe el favorito, lo eliminamos (desmarcar favorito)
            favorite.delete()
            return JsonResponse({"status": "removed"})

        # Si no existe, lo creamos (marcar como favorito)
        return JsonResponse({"status": "added"})

    return HttpResponseBadRequest("Invalid request")


def category_detail(request, pk):
    """
    Vista que muestra el detalle de una categoría específica.

    Solo los usuarios autenticados y con el permiso `MANEJAR_CATEGORIAS` pueden
    acceder a esta vista. Si no se cumplen las condiciones, se redirige al
    usuario a la página de login o a la página de acceso prohibido.

    Args:
        request (HttpRequest): La solicitud HTTP.
        pk (int): El ID de la categoría.

    Returns:
        HttpResponse: Renderiza la plantilla 'article/category_detail.html' o redirige.
    """

    if not request.user.is_authenticated:
        return redirect("login")

    can_manage_categories = request.user.tiene_permisos(
        [PermissionEnum.MANEJAR_CATEGORIAS]
    )

    if not (
        request.user.tiene_permisos([PermissionEnum.MANEJAR_CATEGORIAS])
        or request.user.tiene_permisos([PermissionEnum.VER_CATEGORIAS])
    ):
        return redirect("forbidden")

    category = get_object_or_404(Category, pk=pk)
    return render(
        request,
        "article/category_detail.html",
        {"category": category, "can_manage_categories": can_manage_categories},
    )


def category_create(request):
    """
    Vista que permite la creación de una nueva categoría.

    Solo los usuarios autenticados y con el permiso `MANEJAR_CATEGORIAS` pueden
    acceder a esta vista. Si no se cumplen las condiciones, se redirige al
    usuario a la página de login o a la página de acceso prohibido.

    Args:
        request (HttpRequest): La solicitud HTTP.

    Returns:
        HttpResponse: Renderiza la plantilla 'article/category_form.html' o redirige.
    """

    if not request.user.is_authenticated:
        return redirect("login")

    if not request.user.tiene_permisos([PermissionEnum.MANEJAR_CATEGORIAS]):
        return redirect("forbidden")

    if request.method == "POST":
        form = CategoryForm(request.POST)
        if form.is_valid():
            # Crear una instancia de Category sin guardarla aún
            category = form.save(commit=False)
            # Asigna el ID del usuario al campo createdBy
            category.createdBy = request.user.id
            # Guarda la instancia en la base de datos
            category.save()
            return redirect("category-list")
    else:
        form = CategoryForm()

    return render(request, "article/category_form.html", {"form": form})

    """
    if request.method == "POST":
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("category-list")
    else:
        form = CategoryForm()

    return render(request, "article/category_form.html", {"form": form})"""


def category_update(request, pk):
    """
    Vista que permite actualizar una categoría existente.

    Solo los usuarios autenticados y con el permiso `MANEJAR_CATEGORIAS` pueden
    acceder a esta vista. Si no se cumplen las condiciones, se redirige al
    usuario a la página de login o a la página de acceso prohibido.

    Args:
        request (HttpRequest): La solicitud HTTP.
        pk (int): El ID de la categoría a actualizar.

    Returns:
        HttpResponse: Renderiza la plantilla 'article/category_form.html' o redirige.
    """

    if not request.user.is_authenticated:
        return redirect("login")

    if not request.user.tiene_permisos([PermissionEnum.MANEJAR_CATEGORIAS]):
        return redirect("forbidden")

    category = get_object_or_404(Category, pk=pk)
    if request.method == "POST":
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            return redirect("category-list")
    else:
        form = CategoryForm(instance=category)

    return render(request, "article/category_form.html", {"form": form})


def category_delete(request, pk):
    """
    Vista que permite eliminar una categoría existente.

    Solo los usuarios autenticados y con el permiso `MANEJAR_CATEGORIAS` pueden
    acceder a esta vista. Si no se cumplen las condiciones, se redirige al
    usuario a la página de login o a la página de acceso prohibido.

    Args:
        request (HttpRequest): La solicitud HTTP.
        pk (int): El ID de la categoría a eliminar.

    Returns:
        HttpResponse: Renderiza la plantilla 'article/category_confirm_delete.html' o redirige.
    """

    if not request.user.is_authenticated:
        return redirect("login")

    if not request.user.tiene_permisos([PermissionEnum.MANEJAR_CATEGORIAS]):
        return redirect("forbidden")

    category = get_object_or_404(Category, pk=pk)
    if request.method == "POST":
        category.delete()
        return redirect("category-list")

    return render(
        request, "article/category_confirm_delete.html", {"category": category}
    )


# =============================================================================
# Calificaciones views
# =============================================================================


@login_required
def like_article(request, pk):
    """
    Vista que permite que un usuario marque como 'me gusta' un artículo.

    Si el usuario ya ha marcado el artículo como 'no me gusta', se revierte
    esa acción antes de marcarlo como 'me gusta'. Si es la primera vez que
    marca el artículo, simplemente se incrementa el contador de 'me gusta'.

    Args:
        request (HttpRequest): La solicitud HTTP.
        pk (int): El ID del artículo al que se está aplicando el 'me gusta'.

    Returns:
        HttpResponse: Redirige a la vista de detalle del artículo.
    """
    article = get_object_or_404(Article, pk=pk)
    # Get or create the vote, ensuring that 'vote' is not null
    vote, created = ArticleVote.objects.get_or_create(
        user=request.user, article=article, defaults={"vote": ArticleVote.LIKE}
    )

    if not created and vote.vote != ArticleVote.LIKE:
        # If user previously disliked, undo that dislike
        if vote.vote == ArticleVote.DISLIKE:
            article.dislikes_number -= 1
        # Set the vote to like
        vote.vote = ArticleVote.LIKE
        article.likes_number += 1
        vote.save()

    elif created:
        # If this is the first time the user liked the article
        article.likes_number += 1

    article.save()
    return redirect("article-detail", pk=pk)


@login_required
def dislike_article(request, pk):
    """
    Vista que permite que un usuario marque como 'no me gusta' un artículo.

    Si el usuario ya ha marcado el artículo como 'me gusta', se revierte
    esa acción antes de marcarlo como 'no me gusta'. Si es la primera vez que
    marca el artículo, simplemente se incrementa el contador de 'no me gusta'.

    Args:
        request (HttpRequest): La solicitud HTTP.
        pk (int): El ID del artículo al que se está aplicando el 'no me gusta'.

    Returns:
        HttpResponse: Redirige a la vista de detalle del artículo.
    """
    article = get_object_or_404(Article, pk=pk)
    # Get or create the vote, ensuring that 'vote' is not null
    vote, created = ArticleVote.objects.get_or_create(
        user=request.user, article=article, defaults={"vote": ArticleVote.DISLIKE}
    )

    if not created and vote.vote != ArticleVote.DISLIKE:
        # If user previously liked, undo that like
        if vote.vote == ArticleVote.LIKE:
            article.likes_number -= 1
        # Set the vote to dislike
        vote.vote = ArticleVote.DISLIKE
        article.dislikes_number += 1
        vote.save()

    elif created:
        # If this is the first time the user disliked the article
        article.dislikes_number += 1

    article.save()
    return redirect("article-detail", pk=pk)


def stripe_checkout(request, pk):
    """
    Vista que maneja el proceso de pago con Stripe.

    Verifica si el usuario ya ha completado un pago para la categoría. Si no,
    inicia una sesión de Stripe Checkout para procesar el pago y crea un registro
    de pago pendiente en la base de datos.

    Args:
        request (HttpRequest): La solicitud HTTP.
        pk (int): El ID de la categoría para la que se realiza el pago.

    Returns:
        JsonResponse: Respuesta con el ID de la sesión de Stripe o una señal para redirigir
                      en caso de que el pago ya esté completado.
    """
    try:
        # Verifica si existe un pago completado
        payment_exists = Payment.objects.filter(
            user=request.user, category=pk, status="completed"
        ).exists()

        if payment_exists:
            print("Pago ya existe, redirigiendo a exists.html")  # Verificar condición
            # Retorna una respuesta JSON que indica que ya se realizó el pago
            return JsonResponse({"redirect": "exists"}, status=200)

        # Obtener la categoría o devolver 404 si no existe
        category = get_object_or_404(Category, id=pk)

        # Asignar el precio basado en la categoría
        price_in_cents = int(category.price * 100)  # Convertir a centavos si es necesario

        # Crear el ítem para Stripe Checkout con base en la categoría
        line_items = [
            {
                "price_data": {
                    "currency": "usd",
                    "product_data": {
                        "name": category.name,
                    },
                    "unit_amount": price_in_cents,
                },
                "quantity": 1,
            }
        ]

        # Crear la sesión de Stripe Checkout
        session = stripe.checkout.Session.create(
            payment_method_types=["card"],
            line_items=line_items,
            mode="payment",
            success_url=f'{os.environ.get("URL")}/categories/{category.pk}/success/',
            cancel_url=f'{os.environ.get("URL")}/categories/{category.pk}/cancel/',
        )

        # Crear un nuevo registro de pago con estado 'pending'
        Payment.objects.create(
            user=request.user,
            category=category,
            price=category.price,  # Usar el precio de la categoría
            stripe_payment_id=session.id,
            status="pending",
        )
        
        return JsonResponse({"id": session.id})

    except ObjectDoesNotExist:
        # Si no existe la categoría, devuelve un error 404
        return JsonResponse({"error": "Category does not exist"}, status=404)

    except Exception as e:
        # Manejar cualquier otro error
        return JsonResponse({"error": str(e)}, status=500)


def checkout_page(request, pk):
    """
    Vista que muestra la página de pago para una categoría específica.

    Verifica si el usuario ya ha completado un pago para la categoría y,
    de ser así, redirige a la página de confirmación.

    Args:
        request (HttpRequest): La solicitud HTTP.
        pk (int): El ID de la categoría para la que se realiza el pago.

    Returns:
        HttpResponse: Renderiza la plantilla 'article/checkout.html' o 'article/exists.html'.
    """
    try:
        Payment.objects.filter(
            user=request.user, category=pk, status="completed"
        ).latest("date_paid")
        # Si se encuentra el pago y tiene estado "completed", ir a exists
        return render(request, "article/exists.html")
    except ObjectDoesNotExist:
        return render(
            request,
            "article/checkout.html",
            {
                "STRIPE_PUBLISHABLE_KEY": settings.STRIPE_PUBLIC_KEY,
                "category_id": pk,  # Pasar el ID de la categoría
            },
        )


def payment_success(request, pk):
    """
    Vista que maneja la confirmación de un pago exitoso.

    Verifica el estado del pago, actualiza la base de datos, y envía un correo de
    confirmación al usuario.

    Args:
        request (HttpRequest): La solicitud HTTP.
        pk (int): El ID de la categoría pagada.

    Returns:
        HttpResponse: Renderiza la plantilla 'article/success.html' o 'article/cancel.html'
                      en caso de error.
    """
    category = get_object_or_404(Category, id=pk)
    user = request.user

    try:
        # Recuperar el PaymentIntent desde la base de datos
        payment = Payment.objects.filter(user=user, category=category).latest(
            "date_paid"
        )
        intent = stripe.checkout.Session.retrieve(payment.stripe_payment_id)

        if intent.status == "complete":
            # Actualizar el estado del pago en la base de datos
            payment.status = "completed"
            payment.save()

            # Send confirmation email to the purchaser
            subject = f"Confirmacion de compra por la categoria: {category.name}"
            html_content = f"""
            <h3>Querido {user.username},</h3>
            <p>Gracias por tu compra de:  <strong>{category.name}</strong>.</p>
            <p>La compra fue realizada el:  {payment.date_paid.strftime('%Y-%m-%d %H:%M')}.</p>
            <p>Esperamos disfrutes del contenido!</p>
            """

            # Send the email using the send_email function
            send_email(to=user.email, subject=subject, html=html_content)

            # Redirigir a la página de éxito
            return render(request, "article/success.html", {"category": category})
        else:
            return render(
                request,
                "article/cancel.html",
                {
                    "error": f"El pago no se completó correctamente. Estado: {intent.status}"
                },
            )

    except Payment.DoesNotExist:
        return render(
            request,
            "article/cancel.html",
            {"error": "No se encontró el pago en la base de datos."},
        )


def payment_cancel(request, pk):
    """
    Vista que maneja la cancelación de un pago.

    Marca el pago como cancelado en la base de datos y muestra la página de cancelación.

    Args:
        request (HttpRequest): La solicitud HTTP.
        pk (int): El ID de la categoría relacionada al pago cancelado.

    Returns:
        HttpResponse: Renderiza la plantilla 'article/cancel.html'.
    """
    category = get_object_or_404(Category, id=pk)
    user = request.user
    payment = Payment.objects.filter(user=user, category=category).latest("date_paid")
    payment.status = "cancelled"
    payment.save()
    return render(request, "article/cancel.html")


def category_exists(request, pk):
    """
    Vista que verifica si una categoría específica existe.

    Muestra la página correspondiente si se encuentra la categoría.

    Args:
        request (HttpRequest): La solicitud HTTP.
        pk (int): El ID de la categoría a verificar.

    Returns:
        HttpResponse: Renderiza la plantilla 'article/exists.html'.
    """
    category = get_object_or_404(Category, id=pk)
    user = request.user
    return render(request, "article/exists.html")


@login_required
def sold_categories(request):
    """
    Vista que muestra las categorías vendidas y estadísticas relacionadas.

    Permite filtrar las ventas por rango de fechas y nombre de categoría o usuario.

    Args:
        request (HttpRequest): La solicitud HTTP.

    Returns:
        HttpResponse: Renderiza la plantilla 'article/view_sold_categories.html' o 'article/sold_categories.html'.
    """
    if not request.user.tiene_permisos([PermissionEnum.VER_CATEGORIAS_PAGO]):
        return redirect("forbidden")

    # Get the selected date range from the request (default is 'all')
    date_range = request.GET.get("date_range", "all")
    view_type = request.GET.get("view_type", "default")
    start_date_str = request.GET.get("start_date", None)
    end_date_str = request.GET.get("end_date", None)
    category_name = request.GET.get("category_name", "")
    username = request.GET.get("username", "")

    # Set the filter for the date range
    filter_kwargs = {}
    if date_range == "24h":
        filter_kwargs["date_paid__gte"] = timezone.now() - timedelta(hours=24)
    elif date_range == "7d":
        filter_kwargs["date_paid__gte"] = timezone.now() - timedelta(days=7)
    elif date_range == "30d":
        filter_kwargs["date_paid__gte"] = timezone.now() - timedelta(days=30)
    elif date_range == "365d":
        filter_kwargs["date_paid__gte"] = timezone.now() - timedelta(days=365)

    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
            # Set end_date to include the end of the day (23:59:59)
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d") + timedelta(
                hours=23, minutes=59, seconds=59
            )
            filter_kwargs["date_paid__range"] = (start_date, end_date)
        except ValueError:
            # Handle invalid date format
            return HttpResponseBadRequest("Invalid date format. Use YYYY-MM-DD.")

    # Filter payments based on the selected date range and status 'completed'
    payments = Payment.objects.filter(status="completed", **filter_kwargs)

    if category_name:
        payments = payments.filter(
            category__name__iexact=category_name
        )
    if username:
        payments = payments.filter(
            user__username__iexact=username
        )

    # Group by category name and count the number of payments associated with each category
    categories_sales = (
        payments.values("category__name")
        .annotate(
            total_sales=Count("category"), total_earnings=Sum("price")
        )
        .order_by("-total_sales")
    )

    # Extract category names and corresponding sales for the graph
    categories = [item["category__name"] for item in categories_sales]
    sales = [item["total_sales"] for item in categories_sales]
    earnings = [item["total_earnings"] for item in categories_sales]

    # New variables to track earnings by date for the bar chart
    date_sales = (
        payments.values("date_paid__date")
        .annotate(total_earnings=Sum("price"))
        .order_by("date_paid__date")
    )

    # Extract dates and total earnings for new chart
    dates = [item["date_paid__date"] for item in date_sales]
    dates = [item.strftime("%Y-%m-%d") for item in dates]
    total_earnings_by_date = [item["total_earnings"] for item in date_sales]

    # Get the list of users who bought each category
    buyers_per_category = {
        category["category__name"]: [
            f"{purchase['user__username']} - Costo: ${purchase['price']:.2f} (Fecha: {purchase['date_paid'].strftime('%Y-%m-%d')} Hora: {purchase['date_paid'].strftime('%H:%M:%S')})"
            for purchase in payments.filter(
                category__name=category["category__name"]
            ).values("user__username", "price", "date_paid")
        ]
        for category in categories_sales
    }

    all_categories = Category.objects.filter(payment__status="completed").distinct()
    all_users = CustomUser.objects.filter(payment__status="completed").distinct()

    template_name = (
        "article/view_sold_categories.html"
        if view_type == "list"
        else "article/sold_categories.html"
    )

    # Cálculo del total general de ganancias
    total_general = sum(item["total_earnings"] for item in categories_sales)

    category_data = zip(
        [item["category__name"] for item in categories_sales],
        [item["total_sales"] for item in categories_sales],
        [item["total_earnings"] for item in categories_sales],
        [
            [
                f"{purchase['user__username']} - Costo: ${purchase['price']:.2f} (Fecha: {purchase['date_paid'].strftime('%Y-%m-%d')} Hora: {purchase['date_paid'].strftime('%H:%M:%S')})"
                for purchase in payments.filter(
                    category__name=item["category__name"]
                ).values("user__username", "price", "date_paid")
            ]
            for item in categories_sales
        ],
        ["Tarjeta de crédito"] * len(categories_sales)  # New column value
    )
    
    category_sales_by_date = {}
    for category in categories:
        category_sales_by_date[category] = [
            payments.filter(category__name=category, date_paid__date=date).aggregate(Sum('price'))['price__sum'] or 0
            for date in dates
        ]
        
    detailed_category_data = [
    {
        "category": item["category__name"],
        "buyer": purchase['user__username'],  # Comprador
        "cost": purchase['price'],            # Costo
        "datetime": purchase['date_paid'].strftime('%Y-%m-%d %H:%M:%S'),  # Formatted date and time
        "medio_pago": "Tarjeta de crédito"    # Existing column for payment method
    }
    for item in categories_sales
    for purchase in payments.filter(category__name=item["category__name"]).values("user__username", "price", "date_paid")
]

    return render(
        request,
        template_name,
        {
            "categories": categories,
            "sales": sales,
            "earnings": earnings,
            "dates_json": json.dumps(dates),  # Serialize dates to JSON format
            "total_earnings_by_date_json": json.dumps(total_earnings_by_date),  # Serialize earnings data
            "buyers_per_category": buyers_per_category,
            "category_data": category_data,
            "date_range": date_range,
            "start_date": start_date_str,
            "end_date": end_date_str,
            "total_general": total_general,
            "category_name": category_name,
            "username": username,
            "all_categories": all_categories,
            "all_users": all_users,
            "category_sales_by_date_json": json.dumps(category_sales_by_date),
            "detailed_category_data": detailed_category_data,  # New variable with additional details
        },
    )


@csrf_exempt
@login_required
def sold_categories_suscriptor(request):
    """
    Vista que muestra las categorías vendidas a las que un suscriptor ha accedido.

    Filtra los resultados por fechas y muestra las estadísticas de los pagos realizados.

    Args:
        request (HttpRequest): La solicitud HTTP.

    Returns:
        HttpResponse: Renderiza la plantilla 'article/view_sold_categories_suscriptor.html' o 'article/sold_categories_suscriptor.html'.
    """
    if not request.user.tiene_permisos([PermissionEnum.VER_CATEGORIAS]):
        return redirect("forbidden")

    # Get the selected date range from the request (default is 'all')
    date_range = request.GET.get("date_range", "all")
    view_type = request.GET.get("view_type", "default")
    start_date_str = request.GET.get("start_date", None)
    end_date_str = request.GET.get("end_date", None)

    # Set the filter for the date range
    filter_kwargs = {"user": request.user}

    if date_range == "24h":
        filter_kwargs["date_paid__gte"] = timezone.now() - timedelta(hours=24)
    elif date_range == "7d":
        filter_kwargs["date_paid__gte"] = timezone.now() - timedelta(days=7)
    elif date_range == "30d":
        filter_kwargs["date_paid__gte"] = timezone.now() - timedelta(days=30)
    elif date_range == "365d":
        filter_kwargs["date_paid__gte"] = timezone.now() - timedelta(days=365)

    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
            filter_kwargs["date_paid__range"] = (start_date, end_date)
        except ValueError:
            # Handle invalid date format
            return HttpResponseBadRequest("Invalid date format. Use YYYY-MM-DD.")

    payments = Payment.objects.filter(status="completed", **filter_kwargs).order_by(
        "date_paid"
    )

    # using mayment make a query to get the price and the category
    payments_prices = [payment.price for payment in payments]

    categories_refs = Category.objects.filter(
        pk__in=payments.values_list("category", flat=True)
    ).distinct()

    categories = [category.name for category in categories_refs]

    template_name = (
        "article/view_sold_categories_suscriptor.html"
        if view_type == "list"
        else "article/sold_categories_suscriptor.html"
    )

    total_general = sum(payments_prices)

    return render(
        request,
        template_name,
        {
            "payments": payments,
            "payments_prices": payments_prices,
            "date_range": date_range,  # Pass the selected date range to the template
            "start_date": start_date_str,
            "end_date": end_date_str,
            "categories": categories,
            "total_general": total_general,
        },
    )


@csrf_exempt
@login_required
def download_sold_categories(request):
    """
    View to download an Excel file containing detailed sales information.
    Processes the data from the frontend and generates a file with the formatted details.
    """
    if request.method == "POST":
        data = json.loads(request.body).get("category_data", [])

        # Create an Excel workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas por Categoría"

        # Write headers
        headers = ["Categoría", "Comprador", "Costo", "Fecha y Hora", "Medio de Pago"]
        ws.append(headers)

        # Variable to accumulate total earnings
        total_earnings = 0

        # Add rows based on the detailed_category_data sent from the frontend
        for item in data:
            total_earnings += item['costo']  # Accumulate cost for total earnings
            ws.append([
                item["categoria"],
                item["comprador"],
                f"${item['costo']:.2f}",
                item["fechaHora"],  # Already formatted as 'YYYY-MM-DD HH:MM:SS'
                item["medioPago"]
            ])

        # Add a blank row for separation
        ws.append([])

        # Write total earnings at the end of the sheet
        ws.append(["Total de Ganancias", "", f"${total_earnings:.2f}"])

        # Adjust column widths for better readability
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length

        # Create the response for the Excel file download
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="categorias_vendidas.xlsx"'
        wb.save(response)

        return response
    return JsonResponse({"error": "Invalid request"}, status=400)


@login_required
def download_sold_categories_suscriptor(request):
    """
    Vista que permite a un suscriptor descargar un archivo Excel con detalles de sus compras.

    Args:
        request (HttpRequest): La solicitud HTTP.

    Returns:
        HttpResponse: Respuesta con el archivo Excel para descarga.
    """
    if not request.user.tiene_permisos([PermissionEnum.VER_CATEGORIAS]):
        return redirect("forbidden")

    # Obtener filtros de la solicitud
    start_date_str = request.GET.get("start_date", None)
    end_date_str = request.GET.get("end_date", None)

    # Construir los filtros de usuarios y fechas para la lista de compradores filtrados
    filter_kwargs = {"status": "completed", "user": request.user}
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
            filter_kwargs["date_paid__range"] = (start_date, end_date)
        except ValueError:
            return HttpResponseBadRequest("Invalid date format. Use YYYY-MM-DD.")

    payments = Payment.objects.filter(**filter_kwargs).order_by("date_paid")

    # using mayment make a query to get the price and the category
    payments_prices = [payment.price for payment in payments]

    total_general = sum(payments_prices)

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.title = "Compras por categoría"

    headers = ["Categoría", "Medio de pago", "Precio", "Compradores (Fecha y Hora)"]
    ws.append(headers)

    for payment in payments:
        ws.append(
            [
                payment.category.name,
                "Tarjeta",
                f"${payment.price:.2f}",
                f"{payment.date_paid.strftime('%Y-%m-%d %H:%M:%S')}",
            ]
        )

    ws.append([])
    ws.append(["Total gastado", f"${total_general:.2f}"])

    # Ajustar ancho de columnas para mejor legibilidad
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length

    # Crear la respuesta como archivo Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="categorias_compradas.xlsx"'
    wb.save(response)

    return response

def article_stats(request):
    current_user = request.user
    is_admin = current_user.roles.filter(name="Administrador").exists()
    # Get filter parameters from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    author_id = request.GET.get('author')
    category_id = request.GET.get('category')

    # Base query for published articles
    articles_query = Article.objects.filter(state=ArticleStates.PUBLISHED.value)
    
    if is_admin:  # Check if the user is an admin
        articles_query = Article.objects.filter(state=ArticleStates.PUBLISHED.value)
    else:
        # Non-admin users can only see their own articles
        articles_query = Article.objects.filter(state=ArticleStates.PUBLISHED.value, autor_id=current_user.id)

    # Apply date filter with full day range
    if start_date:
        start_date_parsed = datetime.strptime(start_date, "%Y-%m-%d")
        articles_query = articles_query.filter(published_at__gte=start_date_parsed)
    if end_date:
        end_date_parsed = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(hours=23, minutes=59, seconds=59)
        articles_query = articles_query.filter(published_at__lte=end_date_parsed)

    # Apply author filter
    if author_id:
        articles_query = articles_query.filter(autor_id=author_id)

    # Apply category filter
    if category_id:
        articles_query = articles_query.filter(category_id=category_id)

    # Prepare data for charts (filter articles with likes, dislikes, etc.)
    articles_with_likes = articles_query.filter(likes_number__gt=0).order_by('-likes_number')
    articles_with_dislikes = articles_query.filter(dislikes_number__gt=0).order_by('-dislikes_number')
    articles_with_ratings = articles_query.annotate(avg_rating=Avg('articlevote__rating')).filter(avg_rating__isnull=False).order_by('-avg_rating')
    articles_with_views = articles_query.filter(views_number__gt=0).order_by('-views_number')
    articles_with_shares = articles_query.filter(shares_number__gt=0).order_by('-shares_number')

    # Prepare data dictionaries for charts
    likes_data = {
        "titles": [article.title for article in articles_with_likes],
        "likes": [article.likes_number for article in articles_with_likes],
    }
    dislikes_data = {
        "titles": [article.title for article in articles_with_dislikes],
        "dislikes": [article.dislikes_number for article in articles_with_dislikes],
    }
    avg_rating_data = {
        "titles": [article.title for article in articles_with_ratings],
        "ratings": [article.avg_rating for article in articles_with_ratings],
    }
    avg_views_data = {
        "titles": [article.title for article in articles_with_views],
        "views": [article.views_number for article in articles_with_views],
    }
    shares_data = {
        "titles": [article.title for article in articles_with_shares],
        "shares": [article.shares_number for article in articles_with_shares],
    }

    # Render the template with data and filter options
    return render(request, 'article/article_chart.html', {
        "likes_data": likes_data,
        "dislikes_data": dislikes_data,
        "avg_rating_data": avg_rating_data,
        "avg_views_data": avg_views_data,
        "shares_data": shares_data,
        "authors": CustomUser.objects.all(),  # Assuming CustomUser model is used for authors
        "categories": Category.objects.all(),
        "is_admin": is_admin,
    })

